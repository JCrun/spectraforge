from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import urlsplit, urlunsplit

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font
except ModuleNotFoundError as exc:
    raise SystemExit(
        "缺少依赖 openpyxl，请先执行: pip install openpyxl"
    ) from exc


@dataclass
class GPURecord:
    rank: int
    tier: str
    name: str
    manufacturer: str
    year: int
    detail_url: str
    fp32_tflops: Optional[float]
    pixel_rate_gpixel: Optional[float]
    texture_rate_gtexel: Optional[float]
    memory_size_gb: Optional[float]
    memory_bandwidth_gbs: Optional[float]
    boost_clock_mhz: Optional[float]
    base_clock_mhz: Optional[float]
    release_date: str


def normalize_url(url: str) -> str:
    parsed = urlsplit((url or "").strip())
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path.rstrip("/"), "", ""))


def parse_number_with_unit(text: str, unit_map: Dict[str, float]) -> Optional[float]:
    if not text:
        return None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*([A-Za-z]+)?", text.replace(",", ""))
    if not match:
        return None
    value = float(match.group(1))
    unit = (match.group(2) or "").upper()
    if not unit:
        return value
    for k, factor in unit_map.items():
        if unit.startswith(k):
            return value * factor
    return value


def parse_flops_to_tflops(text: str) -> Optional[float]:
    if not text:
        return None
    cleaned = text.replace(",", "")
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*([TGMK])?FLOPS", cleaned, flags=re.IGNORECASE)
    if not match:
        match = re.search(r"([0-9]+(?:\.[0-9]+)?)", cleaned)
        return float(match.group(1)) if match else None
    value = float(match.group(1))
    unit = (match.group(2) or "T").upper()
    factors = {"T": 1.0, "G": 1e-3, "M": 1e-6, "K": 1e-9}
    return value * factors.get(unit, 1.0)


def detect_tier(rank: int, total: int) -> str:
    ratio = rank / max(total, 1)
    if ratio <= 0.1:
        return "S"
    if ratio <= 0.3:
        return "A"
    if ratio <= 0.6:
        return "B"
    if ratio <= 0.85:
        return "C"
    return "D"


def build_records(payload: Dict[str, object]) -> List[GPURecord]:
    listings = payload.get("listings", [])
    details = payload.get("details", {})

    listing_map: Dict[str, Dict[str, object]] = {}
    for item in listings:
        url = normalize_url(str(item.get("detail_url", "")))
        if url:
            listing_map[url] = item

    rows: List[GPURecord] = []
    for detail_url, detail in details.items():
        norm_url = normalize_url(detail_url)
        listing = listing_map.get(norm_url, {})
        sections = detail.get("sections", {})
        perf = sections.get("Theoretical Performance", {})
        memory = sections.get("Memory", {})
        clocks = sections.get("Clock Speeds", {})
        card = sections.get("Graphics Card", {})
        mobile = sections.get("Mobile Graphics", {})
        igpu = sections.get("Integrated Graphics", {})

        release_date = (
            card.get("Release Date")
            or mobile.get("Release Date")
            or igpu.get("Release Date")
            or ""
        )

        rows.append(
            GPURecord(
                rank=0,
                tier="",
                name=str(listing.get("name") or detail.get("title") or ""),
                manufacturer=str(listing.get("manufacturer") or ""),
                year=int(listing.get("year") or 0),
                detail_url=detail_url,
                fp32_tflops=parse_flops_to_tflops(str(perf.get("FP32 (float)", ""))),
                pixel_rate_gpixel=parse_number_with_unit(str(perf.get("Pixel Rate", "")), {"G": 1.0, "M": 1e-3}),
                texture_rate_gtexel=parse_number_with_unit(
                    str(perf.get("Texture Rate", "")), {"G": 1.0, "M": 1e-3}
                ),
                memory_size_gb=parse_number_with_unit(str(memory.get("Memory Size", "")), {"GB": 1.0, "MB": 1e-3}),
                memory_bandwidth_gbs=parse_number_with_unit(
                    str(memory.get("Memory Bandwidth", "")), {"GB": 1.0, "MB": 1e-3}
                ),
                boost_clock_mhz=parse_number_with_unit(str(clocks.get("Boost Clock", "")), {"MHZ": 1.0, "GHZ": 1000.0}),
                base_clock_mhz=parse_number_with_unit(str(clocks.get("GPU Clock", "")), {"MHZ": 1.0, "GHZ": 1000.0}),
                release_date=release_date,
            )
        )

    rows.sort(key=lambda r: (r.fp32_tflops is None, -(r.fp32_tflops or 0), r.name))
    total = len(rows)
    for idx, row in enumerate(rows, start=1):
        row.rank = idx
        row.tier = detect_tier(idx, total)
    return rows


def write_excel(records: List[GPURecord], output_path: Path, top_n_chart: int = 50) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GPU_Data"

    headers = [
        "Rank",
        "Tier",
        "Name",
        "Manufacturer",
        "Year",
        "FP32_TFLOPS",
        "PixelRate_GPixel_s",
        "TextureRate_GTexel_s",
        "MemorySize_GB",
        "MemoryBandwidth_GB_s",
        "BaseClock_MHz",
        "BoostClock_MHz",
        "ReleaseDate",
        "DetailURL",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for r in records:
        ws.append(
            [
                r.rank,
                r.tier,
                r.name,
                r.manufacturer,
                r.year,
                r.fp32_tflops,
                r.pixel_rate_gpixel,
                r.texture_rate_gtexel,
                r.memory_size_gb,
                r.memory_bandwidth_gbs,
                r.base_clock_mhz,
                r.boost_clock_mhz,
                r.release_date,
                r.detail_url,
            ]
        )

    ladder = wb.create_sheet("Ladder")
    ladder.append(["Rank", "GPU", "FP32_TFLOPS", "Tier"])
    for c in ladder[1]:
        c.font = Font(bold=True)

    top_n = min(top_n_chart, len(records))
    for r in records[:top_n]:
        ladder.append([r.rank, r.name, r.fp32_tflops, r.tier])

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = f"GPU Compute Ladder (Top {top_n} by FP32 TFLOPS)"
    chart.y_axis.title = "GPU"
    chart.x_axis.title = "FP32 (TFLOPS)"
    data = Reference(ladder, min_col=3, min_row=1, max_row=top_n + 1)
    cats = Reference(ladder, min_col=2, min_row=2, max_row=top_n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 18
    chart.width = 30
    ladder.add_chart(chart, "F2")

    write_group_sheet(
        wb,
        title="By_Year",
        records=records,
        group_label="Year",
        group_values=sorted({r.year for r in records if r.year > 0}),
        group_key=lambda r: r.year,
    )
    write_group_sheet(
        wb,
        title="By_Manufacturer",
        records=records,
        group_label="Manufacturer",
        group_values=sorted({r.manufacturer for r in records if r.manufacturer}),
        group_key=lambda r: r.manufacturer,
    )
    write_group_sheet(
        wb,
        title="By_Year_Mfr",
        records=records,
        group_label="Year_Manufacturer",
        group_values=sorted({f"{r.year}_{r.manufacturer}" for r in records if r.year > 0 and r.manufacturer}),
        group_key=lambda r: f"{r.year}_{r.manufacturer}",
    )

    wb.save(output_path)


def write_group_sheet(
    wb: Workbook,
    *,
    title: str,
    records: List[GPURecord],
    group_label: str,
    group_values: List[object],
    group_key,
) -> None:
    ws = wb.create_sheet(title)
    ws.append(
        [
            group_label,
            "GroupRank",
            "GlobalRank",
            "GPU",
            "FP32_TFLOPS",
            "Tier",
            "Manufacturer",
            "Year",
            "DetailURL",
        ]
    )
    for c in ws[1]:
        c.font = Font(bold=True)

    grouped: Dict[object, List[GPURecord]] = {}
    for rec in records:
        key = group_key(rec)
        if key in (None, "", 0):
            continue
        grouped.setdefault(key, []).append(rec)

    for group in group_values:
        items = grouped.get(group, [])
        if not items:
            continue
        items.sort(key=lambda r: (r.fp32_tflops is None, -(r.fp32_tflops or 0), r.name))
        for idx, rec in enumerate(items, start=1):
            ws.append(
                [
                    group,
                    idx,
                    rec.rank,
                    rec.name,
                    rec.fp32_tflops,
                    rec.tier,
                    rec.manufacturer,
                    rec.year,
                    rec.detail_url,
                ]
            )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="将 gpu_specs.json 转为 Excel 并生成天梯图")
    parser.add_argument("--input", default="gpu_specs.json", help="输入 JSON 文件")
    parser.add_argument("--output", default="gpu_ladder.xlsx", help="输出 Excel 文件")
    parser.add_argument("--top-n-chart", type=int, default=50, help="天梯图显示前 N 个 GPU")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    payload = json.loads(Path(args.input).read_text(encoding="utf-8"))
    records = build_records(payload)
    write_excel(records, Path(args.output), top_n_chart=args.top_n_chart)
    print(f"[done] 导出完成: {args.output} (records={len(records)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
